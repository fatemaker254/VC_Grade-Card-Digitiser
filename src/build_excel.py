"""
Stage 4: Excel
Builds one workbook per (semester, subject) group, one row per student,
matching the layout of the user's reference template:
  Roll No | Registration No | Student Name |
  <CODE> Theory | <CODE> Theory Full Marks | <CODE> <Component2> | ... |
  <CODE> Total | <CODE> Grade | <CODE> Credit | <CODE> Credit Points | <CODE> Status |
  ... (repeated per course) ...
  Grand Total | Credits | Credit Points | SGPA | Remarks | Needs Review
Column set per course is derived from whichever components actually appear
in the data (Theoretical is always first, labelled "Theory"; everything
else keeps its printed name, e.g. Practical / Tutorial / Internal Assessment).
"""
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from validate import validate_student


def _component_label(name: str) -> str:
    return "Theory" if name.strip().lower() == "theoretical" else name.strip()


def _num(v):
    if v is None or v == "":
        return None
    try:
        return float(v) if "." in str(v) else int(v)
    except (ValueError, TypeError):
        return v  # keep as text (e.g. "AB", "19 +1") rather than crash


def build_course_schema(records: list) -> list:
    """Returns an ordered list of (course_code, [component_names]) covering
    every course/component seen across all students in this group."""
    schema = {}  # course_code -> {course_name, components: dict preserving order}
    for rec in records:
        for c in rec.get("courses", []):
            code = c.get("course_code") or "UNKNOWN"
            entry = schema.setdefault(code, {"components": {}})
            comps = sorted(c.get("components", []),
                            key=lambda x: 0 if (x.get("name") or "").lower() == "theoretical" else 1)
            for comp in comps:
                entry["components"].setdefault(comp.get("name", "Component"), True)
    return [(code, list(v["components"].keys())) for code, v in schema.items()]


def build_headers(schema: list) -> list:
    headers = ["Roll No", "Registration No", "Student Name"]
    for code, components in schema:
        for comp_name in components:
            label = _component_label(comp_name)
            headers.append(f"{code} {label}")
            headers.append(f"{code} {label} Full Marks")
        headers += [f"{code} Total", f"{code} Grade", f"{code} Credit",
                    f"{code} Credit Points", f"{code} Status"]
    headers += ["Grand Total", "Credits", "Credit Points", "SGPA", "Remarks",
                "Needs Review"]
    return headers


def build_row(record: dict, schema: list) -> list:
    courses_by_code = {c.get("course_code"): c for c in record.get("courses", [])}
    row = [record.get("roll_no", ""), record.get("registration_no", ""),
           record.get("student_name", "")]

    for code, components in schema:
        c = courses_by_code.get(code, {})
        comps_by_name = {comp.get("name"): comp for comp in c.get("components", [])}
        for comp_name in components:
            comp = comps_by_name.get(comp_name, {})
            row.append(_num(comp.get("marks_obtained")))
            row.append(_num(comp.get("full_marks")))
        row += [_num(c.get("total_marks_obtained")), c.get("grade", ""),
                _num(c.get("credit")), _num(c.get("credit_points")), c.get("status", "")]

    validate_notes, _ = validate_student(record, schema)
    review_notes = list(record.get("_review_notes", [])) + validate_notes
    row += [_num(record.get("grand_total_marks_obtained")),
            _num(record.get("grand_total_credit")),
            _num(record.get("grand_total_credit_points")),
            record.get("sgpa", ""), record.get("remarks", ""),
            "; ".join(review_notes) if review_notes else ""]
    return row


def write_subject_workbook(records: list, subject_name: str, out_path: str):
    schema = build_course_schema(records)
    headers = build_headers(schema)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = subject_name[:31]  # Excel sheet-name length limit

    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9E1F2")

    review_col = len(headers)
    for rec in records:
        row = build_row(rec, schema)
        ws.append(row)
        if row[review_col - 1]:
            for c in range(1, len(headers) + 1):
                ws.cell(ws.max_row, c).fill = PatternFill("solid", fgColor="FFF2CC")

    for i, h in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(i)].width = max(10, min(28, len(h) + 2))
    ws.freeze_panes = "D2"

    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
